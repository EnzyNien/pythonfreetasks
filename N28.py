from collections import OrderedDict as OD
import time
import re
import os
try:
	from urlparse import urlparse
except ImportError:
	from urllib.parse import urlparse

import argparse
import bs4 as bs
import cfscrape
import logging
import requests
import csv
from openpyxl import Workbook






class HidemyName():

	type_= OD({'http':'h', 'https':'s', 'Socks 4':'4', 'Socks 5':'5'})
	anonymity_ = OD({'none':'1', 'low':'2', 'middle':'3', 'high':'4'})
	ret_ = OD({'excel':'1', 'csv':'2', 'dict':'3', 'iter':'4'})
	fieldnames = ['ip','port','country','maxtime','type','anon','lastrecently']

	url = 'https://hidemy.name/en/proxy-list/'

	def getRequestHeaders(self):
		return {'accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
				'accept-encoding':'gzip, deflate, wr',
				'accept-language':'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
				'content-type':'application/x-www-form-urlencoded; charset=UTF-8',
				'user-agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.186 Safari/537.36',
				'upgrade-insecure-requests':'1'}

	def getRequestData(self):
		postfix = "?"
		m = 'maxtime={}'.format(self.speed) if self.speed > 0 else ''
		t = 'type={}'.format(self.types) if len(self.types) <4 else ''
		a = 'anon={}'.format(self.anonymity) if len(self.anonymity) <4 else ''
		return postfix + '&'.join([i for i in [m,t,a] if i != ""])

	def compilDict(self,row):
			dict_row = {}
			dict_row['ip'] = row[0]
			dict_row['port'] = row[1]
			dict_row['country'] = row[2]
			dict_row['maxtime'] = row[3]
			dict_row['type'] = row[4]
			dict_row['anon'] = row[5]
			dict_row['lastrecently'] = row[6]
			return dict_row

	def compilInputList(self,data,dict_):
		if not data or not set(data).issubset(set(dict_.keys())):
			return ''.join([dict_[key] for key in dict_.keys()])
		else:
			return ''.join([dict_[key] for key in data])	

	def compilInputInt(self,data,name):
		try:
			data= int(data)
		except:
			err = 'error converting {} string to number'.format(name)
			TypeError(err)
			return	None
		else:
			return data

	def compilResultData(self):
		if not self.result:
			return
		if not self.return_type:
			logging.error('return type is empty'.format(page_index,idx))
			return

		self.return_type = int(self.return_type)

		if self.lastrecently is not None:
			self.result = list(filter(lambda row: row[6] <= self.lastrecently, self.result))
		
		if self.return_type == 1:

			wb = Workbook()
			sheet = wb.active
			for idx,head in enumerate(HidemyName.fieldnames,1):
				sheet.cell(row=1,column=idx).value = head
			for idx,row in enumerate(self.result,2):
				sheet.cell(row=idx, column=1).value = row[0]
				sheet.cell(row=idx, column=2).value = row[1]
				sheet.cell(row=idx, column=3).value = row[2]
				sheet.cell(row=idx, column=4).value = row[3]
				sheet.cell(row=idx, column=5).value = row[4]
				sheet.cell(row=idx, column=6).value = row[5]
				sheet.cell(row=idx, column=7).value = row[6]
			fullpath = os.path.join(self.path,'result.xlsx')
			wb.save(fullpath)

		elif self.return_type == 2:
			
			fullpath = os.path.join(self.path,'result.csv')
			with open(fullpath, 'w', newline='', encoding='utf-8') as csvfile:
				writer = csv.DictWriter(csvfile, fieldnames=HidemyName.fieldnames, delimiter=';')
				writer.writeheader()
				for i in self.result:
					writer.writerow(self.compilDict(i))

		elif self.return_type == 3:

			rerult_dict = {}
			for i in self.result:
				dict_row = self.compilDict(i)
				rerult_dict.update({i[0]:dict_row})
			return rerult_dict

		elif self.return_type == 4:

			rerult_list = []
			for i in self.result:
				rerult_list.append(self.compilDict(i))
			return iter(rerult_list)

		return None

	def connect(self, session = None, url = None, start = None):

		Session = requests.session() if session is None else session
		url = HidemyName.url if url is None else url

		Session.headers = self.getRequestHeaders()
		postfix = self.getRequestData()
		if start is not None:
			postfix += '&start={}'.format(start)
		url += postfix
		scraper = cfscrape.create_scraper(Session,delay=5)

		resp = scraper.request('get',url,headers=scraper.headers)
		return bs.BeautifulSoup(resp.text,'lxml')

	def pars(self,soup,page_index=1):
		table = soup.find('table',{'class':'proxy__t'})
		table = table.tbody
		rows = table.find_all('tr')	
		for idx, row in enumerate(rows,start=1):
			try:
				ip_class = row.find('td',{'class':'tdl'})
				ip = ip_class.get_text(strip=True)
				port, country, maxtime, type_, anon, lastrecently = [i.get_text(strip=True) for i in ip_class.next_siblings]
				maxtime = int(maxtime.split(' ')[0])
				lastrecently = int(lastrecently.split(' ')[0])
				self.result.append([ip,port,country,maxtime,type_,anon,lastrecently])
				logging.info('page {}. row {} is write'.format(page_index,idx))
			except:
				logging.error('page {}. row {} is not write'.format(page_index,idx))
				
	def getPages(self):
		self.result.clear()
		pagelist = []
		soup = self.connect()
		pagination = soup.find('div',{'class':'proxy__pagination'})
		pagination = pagination.ul
		pages = pagination.find_all('li')
		if len(pages) > 1:
			#find max page
			endpage = pages[-1]
			endpage_number = int(endpage.text)
			#get max page number
			endpage_link = endpage.a.get('href',None)
			if endpage_link is None:
				raise ValueError('error find endpage in pagination')
				return
			endslice = self.pattern.search(endpage_link)
			groupdict = endslice.groupdict()
			#find max page slice
			endslice = groupdict.get('slice',None)
			if endslice is None:
				raise ValueError('error find endpage slice')
				return
			global_slice = int(int(endslice)/(endpage_number-1))
			#create slice list
			pagelist = [None] + [int(global_slice)*i for i in range(1,endpage_number)]
		else:
			pagelist = [None]
		for idx, page in enumerate(pagelist,1):
			time.sleep(1)
			#first page if already load
			if page is None:
				self.pars(soup,idx)	
			else:
				soup = self.connect(start=page)
				self.pars(soup,idx)
		return self.compilResultData()

	@staticmethod
	def argParse():

		parser = argparse.ArgumentParser(description='Parsing settings.')
		parser.add_argument("-types","-t", type = str, help ="proxy type: http, https, Socks 4, Socks 5")
		parser.add_argument("-speed","-s", type = str, help ="maximum server delay, ms")
		parser.add_argument("-anonymity","-a", type = str, help ="type of anonymity: none, low, middle, high")
		parser.add_argument("-lastrecently","-lr", type = str, help ="maximum time the server was last refreshed, sec")	
		parser.add_argument("-ret","-r", type = str, help ="type of returned data: excel, csv, dict, iter")

		args= parser.parse_args()
		types = [i.strip() for i in args.types.split(',')] if args.types is not None else []
		anonymity = [i.strip() for i in args.anonymity.split(',')] if args.anonymity is not None else []
		ret = [i.strip() for i in args.ret.split(',')][0] if args.ret is not None else []
		return HidemyName(types,args.speed,anonymity,args.lastrecently,ret)

	def __init__(self,types=None,speed=0,anonymity=None,lastrecently=None,ret=['dict']):
		
		self.result = []

		self.types = self.compilInputList(types,HidemyName.type_)
		self.anonymity = self.compilInputList(anonymity,HidemyName.anonymity_)
		self.return_type = self.compilInputList(ret,HidemyName.ret_)

		self.speed = self.compilInputInt(speed,'speed')
		self.lastrecently = self.compilInputInt(lastrecently,'lastrecently')

		self.pattern = re.compile(r'.+start=(?P<slice>[0-9]*)#')
		self.path = os.path.dirname(__file__)
		fullpath = os.path.join(self.path,'logs.log')
		with open(fullpath,'w'):
			pass
		logging.basicConfig(format = '%(levelname)-8s [%(asctime)s] %(message)s', level = logging.DEBUG, filename = fullpath)



if __name__ == '__main__':
	
	#P = HidemyName.argParse()
	P = HidemyName(types=['http'],speed=800,lastrecently=12,ret=['csv'])
	result = P.getPages()
